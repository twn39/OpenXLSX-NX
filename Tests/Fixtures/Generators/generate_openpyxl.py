import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "TestSheet"

# Write basic values
ws["A1"] = "Hello from openpyxl"
ws["A2"] = 42
ws["A3"] = 3.1415

# Write styled values
ws["B1"] = "Styled"
ws["B1"].font = Font(bold=True, color="FF0000")
ws["B1"].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

wb.save("../openpyxl_generated.xlsx")
